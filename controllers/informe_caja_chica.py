# controllers/informe_caja_chica.py

from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QLabel, 
                             QDateEdit, QPushButton, QTableWidget, QTableWidgetItem,
                             QHeaderView, QGroupBox, QFileDialog, QMessageBox)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QIcon
from sqlalchemy.orm import aliased
import datetime

from models.caja_chica import CajaChicaSesion
from models.usuario import Usuario

try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    LIBRERIAS_OK = True
except ImportError:
    LIBRERIAS_OK = False

class InformeCajaChicaForm(QDialog):
    def __init__(self, session, parent=None):
        super().__init__(parent)
        self.session = session
        self.sesiones_cargadas = []

        self.setWindowTitle("Informe de Caja Chica")
        self.setMinimumSize(1000, 750)
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)
        filtro_group = QGroupBox("Filtros")
        
        main_filter_layout = QVBoxLayout()
        form_layout = QFormLayout()
        self.fecha_desde = QDateEdit(QDate.currentDate().addMonths(-1))
        self.fecha_desde.setCalendarPopup(True)
        self.fecha_hasta = QDateEdit(QDate.currentDate())
        self.fecha_hasta.setCalendarPopup(True)
        form_layout.addRow(QLabel("Desde:"), self.fecha_desde)
        form_layout.addRow(QLabel("Hasta:"), self.fecha_hasta)
        
        button_layout = QHBoxLayout()
        self.btn_generar = QPushButton("Generar Informe")
        self.btn_exportar_pdf = QPushButton(QIcon("imagenes/pdf.jpg"), " Exportar a PDF")
        self.btn_exportar_excel = QPushButton(QIcon("imagenes/excel.png"), " Exportar a Excel")

        button_style = "padding: 8px 15px; font-size: 14px;"
        self.btn_generar.setStyleSheet(button_style)
        self.btn_exportar_pdf.setStyleSheet(button_style)
        self.btn_exportar_excel.setStyleSheet(button_style)
        self.btn_exportar_pdf.setIconSize(self.btn_exportar_pdf.sizeHint() / 2)
        self.btn_exportar_excel.setIconSize(self.btn_exportar_excel.sizeHint() / 2)
        
        button_layout.addWidget(self.btn_generar)
        button_layout.addStretch()
        button_layout.addWidget(self.btn_exportar_pdf)
        button_layout.addWidget(self.btn_exportar_excel)
        
        main_filter_layout.addLayout(form_layout)
        main_filter_layout.addLayout(button_layout)
        filtro_group.setLayout(main_filter_layout)
        
        self.btn_generar.clicked.connect(self.generar_informe)
        self.btn_exportar_pdf.clicked.connect(self.exportar_a_pdf)
        self.btn_exportar_excel.clicked.connect(self.exportar_a_excel)

        if not LIBRERIAS_OK:
            self.btn_exportar_pdf.setEnabled(False)
            self.btn_exportar_excel.setEnabled(False)
            self.btn_exportar_pdf.setToolTip("Librería 'reportlab' no encontrada. Instálala con: pip install reportlab")
            self.btn_exportar_excel.setToolTip("Librería 'openpyxl' no encontrada. Instálala con: pip install openpyxl")
        
        sesiones_group = QGroupBox("Sesiones de Caja")
        sesiones_layout = QVBoxLayout()
        self.tabla_sesiones = QTableWidget()
        self.tabla_sesiones.setColumnCount(7)
        self.tabla_sesiones.setHorizontalHeaderLabels(["ID", "Fecha Apertura", "Fecha Cierre", "Monto Inicial", "Monto Final", "Diferencia", "Usuario Apertura"])
        self.tabla_sesiones.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tabla_sesiones.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabla_sesiones.itemSelectionChanged.connect(self.actualizar_movimientos)
        sesiones_layout.addWidget(self.tabla_sesiones)
        sesiones_group.setLayout(sesiones_layout)

        movimientos_group = QGroupBox("Movimientos de la Sesión Seleccionada")
        movimientos_layout = QVBoxLayout()
        self.tabla_movimientos = QTableWidget()
        self.tabla_movimientos.setColumnCount(5)
        self.tabla_movimientos.setHorizontalHeaderLabels(["ID Mov.", "Fecha y Hora", "Descripción", "Tipo", "Monto"])
        self.tabla_movimientos.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        movimientos_layout.addWidget(self.tabla_movimientos)
        movimientos_group.setLayout(movimientos_layout)

        layout.addWidget(filtro_group)
        layout.addWidget(sesiones_group)
        layout.addWidget(movimientos_group)
        
    def generar_informe(self):
        # (Sin cambios en esta función)
        fecha_desde = self.fecha_desde.date().toPyDate()
        fecha_hasta = self.fecha_hasta.date().toPyDate()
        UsuarioApertura = aliased(Usuario)
        self.sesiones_cargadas = self.session.query(CajaChicaSesion, UsuarioApertura.usuario).\
            join(UsuarioApertura, CajaChicaSesion.idusuario_apertura == UsuarioApertura.idusuario).\
            filter(CajaChicaSesion.fecha_apertura.between(fecha_desde, fecha_hasta)).\
            order_by(CajaChicaSesion.fecha_apertura.desc()).all()
        self.popular_tabla_sesiones()
        self.tabla_movimientos.setRowCount(0)

    def popular_tabla_sesiones(self):
        # (Sin cambios en esta función)
        self.tabla_sesiones.setRowCount(0)
        for sesion, usuario_apertura in self.sesiones_cargadas:
            row = self.tabla_sesiones.rowCount()
            self.tabla_sesiones.insertRow(row)
            item_id = QTableWidgetItem(str(sesion.idcajachica))
            item_id.setData(Qt.UserRole, sesion)
            monto_inicial = f"Gs. {int(sesion.monto_inicial):,}".replace(",", ".")
            monto_final = f"Gs. {int(sesion.monto_final_real or 0):,}".replace(",", ".")
            diferencia = f"Gs. {int(sesion.diferencia or 0):,}".replace(",", ".")
            self.tabla_sesiones.setItem(row, 0, item_id)
            self.tabla_sesiones.setItem(row, 1, QTableWidgetItem(sesion.fecha_apertura.strftime('%d/%m/%Y %H:%M')))
            self.tabla_sesiones.setItem(row, 2, QTableWidgetItem(sesion.fecha_cierre.strftime('%d/%m/%Y %H:%M') if sesion.fecha_cierre else "ABIERTA"))
            self.tabla_sesiones.setItem(row, 3, QTableWidgetItem(monto_inicial))
            self.tabla_sesiones.setItem(row, 4, QTableWidgetItem(monto_final))
            self.tabla_sesiones.setItem(row, 5, QTableWidgetItem(diferencia))
            self.tabla_sesiones.setItem(row, 6, QTableWidgetItem(usuario_apertura))
        self.tabla_sesiones.resizeColumnsToContents()

    def actualizar_movimientos(self):
        # (Sin cambios en esta función)
        selected_items = self.tabla_sesiones.selectedItems()
        if not selected_items: return
        self.tabla_movimientos.setRowCount(0)
        selected_row = selected_items[0].row()
        sesion_obj = self.tabla_sesiones.item(selected_row, 0).data(Qt.UserRole)
        if not sesion_obj: return
        for mov in sesion_obj.movimientos:
            row = self.tabla_movimientos.rowCount()
            self.tabla_movimientos.insertRow(row)
            monto_formateado = f"Gs. {int(mov.monto):,}".replace(",", ".")
            monto_item = QTableWidgetItem(monto_formateado)
            monto_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tabla_movimientos.setItem(row, 0, QTableWidgetItem(str(mov.idmovimiento)))
            self.tabla_movimientos.setItem(row, 1, QTableWidgetItem(mov.fecha_hora.strftime('%d/%m/%Y %H:%M')))
            self.tabla_movimientos.setItem(row, 2, QTableWidgetItem(mov.descripcion))
            self.tabla_movimientos.setItem(row, 3, QTableWidgetItem(mov.tipo_movimiento))
            self.tabla_movimientos.setItem(row, 4, monto_item)
        self.tabla_movimientos.resizeColumnsToContents()
        self.tabla_movimientos.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)

    def exportar_a_pdf(self):
        if not self.sesiones_cargadas:
            QMessageBox.warning(self, "Sin Datos", "Debe generar un informe primero.")
            return

        timestamp = datetime.datetime.now().strftime("%d%m%y_%H%M")
        nombre_sugerido = f"inf_caja_chica_{timestamp}.pdf"
        filename, _ = QFileDialog.getSaveFileName(self, "Guardar PDF", nombre_sugerido, "Archivos PDF (*.pdf)")
        if not filename: return

        doc = SimpleDocTemplate(filename, pagesize=landscape(letter))
        story, styles = [], getSampleStyleSheet()
        story.append(Paragraph("Informe de Caja Chica", styles['Title']))
        story.append(Paragraph(f"Periodo: {self.fecha_desde.text()} al {self.fecha_hasta.text()}", styles['Normal']))
        story.append(Spacer(1, 12))

        for sesion, usuario in self.sesiones_cargadas:
            story.append(Paragraph(f"<b>Sesión ID: {sesion.idcajachica}</b> - Abierta por: {usuario} el {sesion.fecha_apertura.strftime('%d/%m/%Y %H:%M')}", styles['h3']))
            
            # <<< CAMBIO: Cálculo de totales antes de crear la tabla >>>
            total_movimientos = 0
            mov_data = [["ID", "Fecha/Hora", "Descripción", "Tipo", "Monto"]]
            for mov in sesion.movimientos:
                total_movimientos += mov.monto
                mov_data.append([
                    str(mov.idmovimiento), mov.fecha_hora.strftime('%d/%m/%Y %H:%M'),
                    Paragraph(mov.descripcion, styles['Normal']), mov.tipo_movimiento,
                    f"Gs. {int(mov.monto):,}".replace(",", ".")
                ])
            
            # <<< CAMBIO: Agregamos las filas de totales al final de los datos >>>
            saldo_final = sesion.monto_inicial - total_movimientos
            mov_data.append(['', '', '', '', '']) # Separador
            mov_data.append(['', '', '', 'Total Pagado:', f"Gs. {int(total_movimientos):,}".replace(",", ".")])
            mov_data.append(['', '', '', 'Monto Inicial:', f"Gs. {int(sesion.monto_inicial):,}".replace(",", ".")])
            mov_data.append(['', '', '', 'Saldo Final:', f"Gs. {int(saldo_final):,}".replace(",", ".")])

            t = Table(mov_data, colWidths=[40, 90, 350, 80, 80])
            # <<< CAMBIO: Agregamos estilos para las filas de totales >>>
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -5), colors.beige), # Hasta antes de los totales
                ('GRID', (0, 0), (-1, -5), 1, colors.black), # Grilla hasta antes de los totales
                ('LINEABOVE', (3, -3), (-1, -3), 1, colors.black), # Línea sobre los totales
                ('ALIGN', (3, -3), (3, -1), 'RIGHT'), # Alineación de etiquetas de total
                ('FONTNAME', (3, -3), (3, -1), 'Helvetica-Bold'), # Negrita para etiquetas de total
            ]))
            story.append(t)
            story.append(Spacer(1, 24))

        doc.build(story)
        QMessageBox.information(self, "Éxito", f"Informe guardado como PDF en:\n{filename}")

    def exportar_a_excel(self):
        if not self.sesiones_cargadas:
            QMessageBox.warning(self, "Sin Datos", "Debe generar un informe primero.")
            return

        timestamp = datetime.datetime.now().strftime("%d%m%y_%H%M")
        nombre_sugerido = f"inf_caja_chica_{timestamp}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", nombre_sugerido, "Archivos Excel (*.xlsx)")
        if not filename: return

        wb = Workbook(); ws = wb.active
        ws.title = "Informe Caja Chica"
        ws.append(["Informe de Caja Chica"])
        ws.append([f"Periodo: {self.fecha_desde.text()} al {self.fecha_hasta.text()}"])
        ws.append([])
        header_font = Font(bold=True)
        bold_font = Font(bold=True)
        
        for sesion, usuario in self.sesiones_cargadas:
            ws.append([f"Sesión ID: {sesion.idcajachica}", f"Usuario Apertura: {usuario}", f"Fecha Apertura: {sesion.fecha_apertura.strftime('%d/%m/%Y %H:%M')}"])
            mov_header = ["ID Mov.", "Fecha y Hora", "Descripción", "Tipo", "Monto"]
            ws.append(mov_header)
            for cell in ws[ws.max_row]: cell.font = header_font

            # <<< CAMBIO: Calculamos totales mientras agregamos filas >>>
            total_movimientos = 0
            for mov in sesion.movimientos:
                total_movimientos += mov.monto
                ws.append([
                    mov.idmovimiento, mov.fecha_hora.strftime('%d/%m/%Y %H:%M'),
                    mov.descripcion, mov.tipo_movimiento, int(mov.monto)
                ])
                # Formato de número para la celda del monto
                ws.cell(row=ws.max_row, column=5).number_format = '#,##0'

            # <<< CAMBIO: Agregamos las filas de totales >>>
            saldo_final = sesion.monto_inicial - total_movimientos
            ws.append([]) # Separador
            
            ws.append(['', '', '', 'Total Pagado:', int(total_movimientos)])
            ws.cell(row=ws.max_row, column=4).font = bold_font
            ws.cell(row=ws.max_row, column=5).number_format = '#,##0'
            
            ws.append(['', '', '', 'Monto Inicial:', int(sesion.monto_inicial)])
            ws.cell(row=ws.max_row, column=4).font = bold_font
            ws.cell(row=ws.max_row, column=5).number_format = '#,##0'
            
            ws.append(['', '', '', 'Saldo Final:', int(saldo_final)])
            ws.cell(row=ws.max_row, column=4).font = bold_font
            ws.cell(row=ws.max_row, column=5).number_format = '#,##0'

            ws.append([])

        # Autoajuste de columnas
        for i, column_cells in enumerate(ws.columns):
            max_length = 0
            column = get_column_letter(i + 1)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filename)
        QMessageBox.information(self, "Éxito", f"Informe guardado como Excel en:\n{filename}")